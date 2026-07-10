import io
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from ..core import get_db
from ..models import Student
from ..schemas import StudentCreate, StudentUpdate, StudentOut, Statistics
from ..services import get_active_user, get_admin_user, get_active_user_or_apikey, get_admin_user_or_apikey
from ..models import User

router = APIRouter(prefix="/students", tags=["Students"])


def _get_or_404(db: Session, student_id: int) -> Student:
    s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    return s


@router.get("/", response_model=List[StudentOut])
def list_students(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    q = db.query(Student)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(
            Student.full_name.ilike(like),
            Student.student_code.ilike(like),
            Student.email.ilike(like),
            Student.class_name.ilike(like),
        ))
    if status:
        q = q.filter(Student.status == status)
    return q.order_by(Student.id).offset(skip).limit(limit).all()


@router.get("/statistics", response_model=Statistics)
def get_statistics(
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    total = db.query(Student).count()
    active = db.query(Student).filter(Student.status == "active").count()
    inactive = db.query(Student).filter(Student.status == "inactive").count()
    graduated = db.query(Student).filter(Student.status == "graduated").count()
    avg = db.query(func.avg(Student.gpa)).scalar() or 0.0
    return Statistics(total=total, active=active, inactive=inactive, graduated=graduated, average_gpa=round(float(avg), 2))


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    students = db.query(Student).order_by(Student.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách học sinh"

    headers = ["Mã số", "Họ tên", "Ngày sinh", "Giới tính", "Email", "Điện thoại", "Lớp", "Điểm TB", "Trạng thái"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_map = {"active": "Đang học", "inactive": "Tạm nghỉ", "graduated": "Đã tốt nghiệp"}
    for row, s in enumerate(students, 2):
        ws.append([
            s.student_code, s.full_name,
            s.date_of_birth.strftime("%d/%m/%Y") if s.date_of_birth else "",
            s.gender, s.email, s.phone or "",
            s.class_name or "", s.gpa or 0,
            status_map.get(s.status, s.status),
        ])

    col_widths = [12, 25, 12, 10, 30, 14, 10, 10, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"},
    )


@router.get("/export/pdf")
def export_pdf(
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed")

    students = db.query(Student).order_by(Student.id).all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, spaceAfter=10, alignment=1)

    status_map = {"active": "Đang học", "inactive": "Tạm nghỉ", "graduated": "Đã tốt nghiệp"}
    data = [["Mã số", "Họ tên", "Ngày sinh", "Giới tính", "Email", "Lớp", "Điểm TB", "Trạng thái"]]
    for s in students:
        data.append([
            s.student_code, s.full_name,
            s.date_of_birth.strftime("%d/%m/%Y") if s.date_of_birth else "",
            s.gender, s.email, s.class_name or "",
            f"{s.gpa:.1f}" if s.gpa else "0.0",
            status_map.get(s.status, s.status),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E0")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))

    elements = [Paragraph("DANH SÁCH HỌC SINH", title_style), Spacer(1, 10), table]
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=students.pdf"},
    )


@router.get("/import/template")
def download_import_template(
    _: User = Depends(get_active_user_or_apikey),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Học sinh"

    headers = ["Mã số", "Họ tên", "Ngày sinh (dd/mm/yyyy)", "Giới tính", "Email", "Điện thoại", "Lớp", "Điểm TB", "Trạng thái"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # One example row to guide the user
    ws.append(["HS001", "Nguyễn Văn A", "01/09/2008", "Nam", "vana@example.com", "0901234567", "10A1", 8.5, "active"])

    col_widths = [12, 25, 20, 10, 30, 14, 10, 10, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Dropdown validation for Gender and Status to reduce input errors
    gender_dv = DataValidation(type="list", formula1='"Nam,Nữ,Khác"', allow_blank=False)
    status_dv = DataValidation(type="list", formula1='"active,inactive,graduated"', allow_blank=False)
    ws.add_data_validation(gender_dv)
    ws.add_data_validation(status_dv)
    gender_dv.add("D2:D1000")
    status_dv.add("I2:I1000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"},
    )


@router.post("/import/excel", response_model=List[StudentOut], status_code=status.HTTP_201_CREATED)
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File phải có định dạng .xlsx hoặc .xls")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Không thể đọc file Excel. File có thể bị hỏng.")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]

    if not rows:
        raise HTTPException(status_code=400, detail="File không có dữ liệu để import.")

    valid_genders = {"Nam", "Nữ", "Khác"}
    valid_statuses = {"active", "inactive", "graduated"}

    parsed: List[dict] = []
    errors: List[str] = []
    codes_seen: dict[str, int] = {}
    emails_seen: dict[str, int] = {}

    for idx, row in enumerate(rows, start=2):  # row 1 is header
        row = list(row) + [None] * (9 - len(row))
        student_code, full_name, dob_raw, gender, email, phone, class_name, gpa, status_val = row[:9]

        row_errors = []

        student_code = str(student_code).strip() if student_code else ""
        full_name = str(full_name).strip() if full_name else ""
        gender = str(gender).strip() if gender else ""
        email = str(email).strip() if email else ""
        phone = str(phone).strip() if phone else None
        class_name = str(class_name).strip() if class_name else None
        status_val = str(status_val).strip() if status_val else "active"

        if not student_code:
            row_errors.append("thiếu mã số")
        elif student_code in codes_seen:
            row_errors.append(f"mã số trùng với dòng {codes_seen[student_code]}")
        else:
            codes_seen[student_code] = idx

        if not full_name:
            row_errors.append("thiếu họ tên")

        # Parse date of birth (accept datetime object or dd/mm/yyyy string)
        dob = None
        if dob_raw is None or dob_raw == "":
            row_errors.append("thiếu ngày sinh")
        elif hasattr(dob_raw, "date") and callable(getattr(dob_raw, "date")):
            try:
                dob = dob_raw.date()
            except (AttributeError, TypeError):
                dob = dob_raw
        else:
            try:
                dob = datetime.strptime(str(dob_raw).strip(), "%d/%m/%Y").date()
            except ValueError:
                row_errors.append("ngày sinh sai định dạng (cần dd/mm/yyyy)")

        if gender not in valid_genders:
            row_errors.append(f"giới tính không hợp lệ ({gender or 'trống'})")

        if not email or "@" not in email:
            row_errors.append("email không hợp lệ")
        elif email in emails_seen:
            row_errors.append(f"email trùng với dòng {emails_seen[email]}")
        else:
            emails_seen[email] = idx

        try:
            gpa_val = float(gpa) if gpa not in (None, "") else 0.0
            if not (0.0 <= gpa_val <= 10.0):
                row_errors.append("điểm TB phải từ 0 đến 10")
        except (ValueError, TypeError):
            row_errors.append("điểm TB không hợp lệ")
            gpa_val = 0.0

        if status_val not in valid_statuses:
            row_errors.append(f"trạng thái không hợp lệ ({status_val})")

        if row_errors:
            errors.append(f"Dòng {idx}: {', '.join(row_errors)}")
            continue

        parsed.append(dict(
            student_code=student_code, full_name=full_name, date_of_birth=dob,
            gender=gender, email=email, phone=phone, class_name=class_name,
            gpa=gpa_val, status=status_val,
        ))

    # Cross-check against existing DB records (still part of "fail if any row is invalid")
    existing_codes = {c for (c,) in db.query(Student.student_code).all()}
    existing_emails = {e for (e,) in db.query(Student.email).all()}

    for item in parsed:
        if item["student_code"] in existing_codes:
            errors.append(f"Mã số {item['student_code']} đã tồn tại trong hệ thống")
        if item["email"] in existing_emails:
            errors.append(f"Email {item['email']} đã tồn tại trong hệ thống")

    if errors:
        raise HTTPException(
            status_code=422,
            detail={
                "message": f"Import thất bại: phát hiện {len(errors)} lỗi. Không có dữ liệu nào được lưu.",
                "errors": errors,
            },
        )

    new_students = [Student(**item) for item in parsed]
    db.add_all(new_students)
    db.commit()
    for s in new_students:
        db.refresh(s)
    return new_students


@router.post("/", response_model=StudentOut, status_code=status.HTTP_201_CREATED)
def create_student(
    payload: StudentCreate,
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    if db.query(Student).filter(Student.student_code == payload.student_code).first():
        raise HTTPException(status_code=400, detail="Student code already exists")
    if db.query(Student).filter(Student.email == payload.email).first():
        raise HTTPException(status_code=400, detail="Email already exists")
    s = Student(**payload.model_dump())
    db.add(s)
    db.commit()
    db.refresh(s)
    return s


@router.get("/{student_id}", response_model=StudentOut)
def get_student(student_id: int, db: Session = Depends(get_db), _: User = Depends(get_active_user_or_apikey)):
    return _get_or_404(db, student_id)


@router.put("/{student_id}", response_model=StudentOut)
def update_student(
    student_id: int,
    payload: StudentUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    s = _get_or_404(db, student_id)
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)
    return s


@router.delete("/{student_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_student(
    student_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user_or_apikey),  # Only admin can delete
):
    s = _get_or_404(db, student_id)
    db.delete(s)
    db.commit()
