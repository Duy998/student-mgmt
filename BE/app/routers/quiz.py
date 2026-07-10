import io
import random
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from ..core import get_db
from ..models import Question, QuizAttempt, User
from ..schemas import QuestionOut, QuestionAdminOut, QuizSubmitPayload, QuizResult, AttemptOut
from ..services import get_active_user, get_admin_user, get_active_user_or_apikey, get_admin_user_or_apikey

router = APIRouter(prefix="/quiz", tags=["Quiz"])

VALID_ANSWERS = {"A", "B", "C", "D"}


# ── Admin: Question Bank ───────────────────────────────────────────────────────

@router.get("/questions", response_model=List[QuestionAdminOut])
def list_questions(
    grade: Optional[int] = Query(None),
    topic: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user_or_apikey),
):
    q = db.query(Question)
    if grade:
        q = q.filter(Question.grade == grade)
    if topic:
        q = q.filter(Question.topic.ilike(f"%{topic}%"))
    if search:
        q = q.filter(or_(
            Question.content.ilike(f"%{search}%"),
            Question.code.ilike(f"%{search}%"),
        ))
    return q.order_by(Question.code).offset(skip).limit(limit).all()


@router.delete("/questions/{question_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_question(
    question_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user_or_apikey),
):
    q = db.query(Question).filter(Question.id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    db.delete(q)
    db.commit()


@router.get("/questions/template")
def download_question_template(_: User = Depends(get_admin_user_or_apikey)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Câu hỏi"

    headers = ["Mã", "Khối", "Chủ đề", "Mức độ", "Loại", "Câu hỏi", "A", "B", "C", "D", "Đáp án"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sample rows matching the user's format
    ws.append(["L6-001", 6, "Thông tin và dữ liệu", "Nhận biết", "TN",
               "Câu 1. Nội dung kiến thức về Thông tin và dữ liệu.",
               "Phương án A", "Phương án B", "Phương án C", "Phương án D", "A"])
    ws.append(["L6-002", 6, "Thiết bị máy tính", "Nhận biết", "TN",
               "Câu 2. Nội dung kiến thức về Thiết bị máy tính.",
               "Phương án A", "Phương án B", "Phương án C", "Phương án D", "B"])

    col_widths = [10, 7, 25, 12, 7, 50, 20, 20, 20, 20, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    answer_dv = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=False)
    ws.add_data_validation(answer_dv)
    answer_dv.add("K2:K5000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=question_template.xlsx"},
    )


@router.post("/questions/import", status_code=status.HTTP_201_CREATED)
async def import_questions(
    file: UploadFile = File(...),
    overwrite: bool = Query(False, description="Overwrite existing codes"),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user_or_apikey),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File phải có định dạng .xlsx hoặc .xls")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Không thể đọc file Excel.")

    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
            if any(c is not None and str(c).strip() for c in r)]

    if not rows:
        raise HTTPException(status_code=400, detail="File không có dữ liệu.")

    errors: List[str] = []
    parsed: List[dict] = []
    codes_seen: dict = {}

    for idx, row in enumerate(rows, start=2):
        row = list(row) + [None] * (11 - len(list(row)))
        code, grade, topic, level, q_type, content_text, a, b, c, d, answer = [
            str(v).strip() if v is not None else "" for v in row[:11]
        ]
        row_errors = []

        if not code:
            row_errors.append("thiếu Mã")
        elif code in codes_seen:
            row_errors.append(f"Mã trùng với dòng {codes_seen[code]}")
        else:
            codes_seen[code] = idx

        if not grade.isdigit():
            row_errors.append("Khối phải là số nguyên")
        if not topic:
            row_errors.append("thiếu Chủ đề")
        if not level:
            row_errors.append("thiếu Mức độ")
        if not content_text:
            row_errors.append("thiếu Câu hỏi")
        if not a or not b or not c or not d:
            row_errors.append("thiếu một hoặc nhiều phương án A/B/C/D")
        if answer.upper() not in VALID_ANSWERS:
            row_errors.append(f"Đáp án '{answer}' không hợp lệ (phải là A/B/C/D)")

        if row_errors:
            errors.append(f"Dòng {idx}: {', '.join(row_errors)}")
        else:
            parsed.append(dict(
                code=code, grade=int(grade), topic=topic, level=level,
                q_type=q_type or "TN", content=content_text,
                option_a=a, option_b=b, option_c=c, option_d=d,
                answer=answer.upper(),
            ))

    if errors:
        raise HTTPException(status_code=422, detail={
            "message": f"Import thất bại: {len(errors)} lỗi. Không có câu hỏi nào được lưu.",
            "errors": errors,
        })

    existing_codes = {q.code for q in db.query(Question.code).all()}
    added, updated = 0, 0
    for item in parsed:
        if item["code"] in existing_codes:
            if overwrite:
                q = db.query(Question).filter(Question.code == item["code"]).first()
                for k, v in item.items():
                    setattr(q, k, v)
                updated += 1
            # else skip silently (code exists, overwrite=False)
        else:
            db.add(Question(**item))
            added += 1

    db.commit()
    return {"added": added, "updated": updated, "skipped": len(parsed) - added - updated}


# ── Student: Draw & Submit Quiz ───────────────────────────────────────────────

@router.get("/draw", response_model=List[QuestionOut])
def draw_questions(
    count: int = Query(10, ge=1, le=100),
    grade: Optional[int] = Query(None),
    topic: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_active_user_or_apikey),
):
    """Return N random questions WITHOUT the answer field."""
    q = db.query(Question)
    if grade:
        q = q.filter(Question.grade == grade)
    if topic:
        q = q.filter(Question.topic.ilike(f"%{topic}%"))
    pool = q.all()
    if not pool:
        raise HTTPException(status_code=404, detail="Không có câu hỏi nào phù hợp.")
    drawn = random.sample(pool, min(count, len(pool)))
    return drawn   # QuestionOut does NOT include answer — schema enforces this


@router.post("/submit", response_model=QuizResult)
def submit_quiz(
    payload: QuizSubmitPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_active_user_or_apikey),
):
    if not payload.answers:
        raise HTTPException(status_code=400, detail="Không có câu trả lời nào.")

    question_ids = [a.question_id for a in payload.answers]
    questions = {q.id: q for q in db.query(Question).filter(Question.id.in_(question_ids)).all()}

    correct = 0
    for a in payload.answers:
        q = questions.get(a.question_id)
        if q and a.chosen.upper() == q.answer:
            correct += 1

    total = len(payload.answers)
    score = round(correct / total * 10, 2) if total else 0.0

    attempt = QuizAttempt(
        user_id=current_user.id,
        username=current_user.username,
        score=score,
        total=total,
        correct=correct,
        time_spent=payload.time_spent,
    )
    db.add(attempt)
    db.commit()

    return QuizResult(score=score, correct=correct, total=total, time_spent=payload.time_spent)


# ── Admin: Results ────────────────────────────────────────────────────────────

@router.get("/results", response_model=List[AttemptOut])
def list_results(
    username: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(get_admin_user_or_apikey),
):
    q = db.query(QuizAttempt)
    if username:
        q = q.filter(QuizAttempt.username.ilike(f"%{username}%"))
    return q.order_by(QuizAttempt.submitted_at.desc()).offset(skip).limit(limit).all()
